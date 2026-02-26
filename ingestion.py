"""
Ingestion Module — RIS / BIB / CSV reference parsing, deduplication,
and title-abstract LLM screening.

Covers PRISMA Phases 2 & 3:
  Phase 2 — Import records from PubMed/Scopus/WoS exports → deduplicate
  Phase 3 — LLM screens titles & abstracts (no PDFs needed at this stage)

Public API
----------
parse_references(path)         → list[dict]  (normalised record dicts)
deduplicate(records)           → (list[dict], DeduplicationStats)
AbstractScreener(llm_manager)  → .screen_all(records, criteria, callback)
export_records_to_excel(records, path)
export_records_to_csv(records, path)
"""

__version__ = "3.1.0"

import re
import csv
import json
import time
import threading
import hashlib
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional, Tuple, Callable, Any
from pathlib import Path
import logging

logger = logging.getLogger("SLR_Ingestion")


# ─────────────────────────────────────────────────────────────────────────────
# Normalised record schema
# ─────────────────────────────────────────────────────────────────────────────

def _empty_record() -> Dict[str, Any]:
    return {
        "record_id":  "",
        "title":      "",
        "abstract":   "",
        "authors":    "",
        "year":       "",
        "journal":    "",
        "doi":        "",
        "keywords":   "",
        "source_file": "",
        # set during screening
        "decision":   "",
        "rationale":  "",
        "human_override": False,
    }


# ─────────────────────────────────────────────────────────────────────────────
# RIS parser
# ─────────────────────────────────────────────────────────────────────────────

def _parse_ris(path: str) -> List[Dict]:
    """Parse .ris file using rispy; fall back to manual parser if unavailable."""
    records = []
    fname = Path(path).name

    try:
        import rispy
        with open(path, encoding="utf-8", errors="replace") as fh:
            entries = rispy.load(fh)
        for e in entries:
            r = _empty_record()
            r["source_file"] = fname
            r["title"]     = e.get("title", e.get("primary_title", ""))
            ab = e.get("abstract", "")
            if isinstance(ab, list):
                ab = " ".join(ab)
            r["abstract"]  = ab
            authors = e.get("authors", e.get("first_authors", []))
            if isinstance(authors, list):
                authors = "; ".join(authors)
            r["authors"]   = authors
            r["year"]      = str(e.get("year", e.get("publication_year", "")))
            r["journal"]   = e.get("journal_name", e.get("secondary_title", ""))
            r["doi"]       = e.get("doi", "")
            kw = e.get("keywords", [])
            r["keywords"]  = "; ".join(kw) if isinstance(kw, list) else str(kw)
            r["record_id"] = _make_id(r)
            if r["title"].strip():
                records.append(r)
    except ImportError:
        logger.warning("rispy not installed — using fallback RIS parser")
        records = _parse_ris_manual(path, fname)
    except Exception as exc:
        logger.error(f"RIS parse error: {exc}")
        records = _parse_ris_manual(path, fname)

    return records


def _parse_ris_manual(path: str, fname: str) -> List[Dict]:
    """Tag-by-tag RIS parser that works without rispy."""
    records = []
    r = _empty_record()
    r["source_file"] = fname
    in_record = False

    with open(path, encoding="utf-8", errors="replace") as fh:
        for line in fh:
            line = line.rstrip("\n")
            if len(line) < 2:
                continue
            tag  = line[:2].strip()
            val  = line[6:].strip() if len(line) > 6 else ""

            if tag == "TY":
                in_record = True
                r = _empty_record()
                r["source_file"] = fname
            elif tag == "ER" and in_record:
                in_record = False
                r["record_id"] = _make_id(r)
                if r["title"].strip():
                    records.append(r)
            elif tag in ("TI", "T1", "CT"):
                r["title"] = val
            elif tag == "AB":
                r["abstract"] += (" " + val if r["abstract"] else val)
            elif tag in ("AU", "A1", "A2"):
                r["authors"] += ("; " + val if r["authors"] else val)
            elif tag in ("PY", "Y1", "DA"):
                r["year"] = val[:4]
            elif tag in ("JO", "JF", "T2", "SO"):
                r["journal"] = val
            elif tag == "DO":
                r["doi"] = val
            elif tag == "KW":
                r["keywords"] += ("; " + val if r["keywords"] else val)

    return records


# ─────────────────────────────────────────────────────────────────────────────
# BIB parser
# ─────────────────────────────────────────────────────────────────────────────

def _parse_bib(path: str) -> List[Dict]:
    """Parse .bib file using bibtexparser; fall back to regex parser."""
    fname = Path(path).name
    records = []

    try:
        import bibtexparser
        with open(path, encoding="utf-8", errors="replace") as fh:
            db = bibtexparser.load(fh)
        for e in db.entries:
            r = _empty_record()
            r["source_file"] = fname
            r["title"]    = _clean_bib(e.get("title", ""))
            r["abstract"] = _clean_bib(e.get("abstract", ""))
            r["authors"]  = _clean_bib(e.get("author", "")).replace(" and ", "; ")
            r["year"]     = e.get("year", "")
            r["journal"]  = _clean_bib(e.get("journal", e.get("booktitle", "")))
            r["doi"]      = e.get("doi", "")
            r["keywords"] = _clean_bib(e.get("keywords", ""))
            r["record_id"] = _make_id(r)
            if r["title"].strip():
                records.append(r)
    except ImportError:
        logger.warning("bibtexparser not installed — using fallback BIB parser")
        records = _parse_bib_regex(path, fname)
    except Exception as exc:
        logger.error(f"BIB parse error: {exc}")
        records = _parse_bib_regex(path, fname)

    return records


def _clean_bib(text: str) -> str:
    """Remove LaTeX brace wrappers like {Some Title}."""
    return re.sub(r'\{([^{}]*)\}', r'\1', text).strip()


def _parse_bib_regex(path: str, fname: str) -> List[Dict]:
    """Minimal regex-based BIB parser."""
    records = []
    with open(path, encoding="utf-8", errors="replace") as fh:
        content = fh.read()

    for block in re.findall(r'@\w+\{[^@]+', content, re.DOTALL):
        r = _empty_record()
        r["source_file"] = fname

        def _get(tag):
            m = re.search(rf'\b{tag}\s*=\s*[{{"](.+?)[{{}}"]\s*[,}}]', block,
                          re.IGNORECASE | re.DOTALL)
            return _clean_bib(m.group(1)) if m else ""

        r["title"]    = _get("title")
        r["abstract"] = _get("abstract")
        r["authors"]  = _get("author").replace(" and ", "; ")
        r["year"]     = _get("year")
        r["journal"]  = _get("journal") or _get("booktitle")
        r["doi"]      = _get("doi")
        r["keywords"] = _get("keywords")
        r["record_id"] = _make_id(r)
        if r["title"].strip():
            records.append(r)

    return records


# ─────────────────────────────────────────────────────────────────────────────
# CSV parser
# ─────────────────────────────────────────────────────────────────────────────

def _parse_csv_file(path: str) -> List[Dict]:
    """
    Parse a CSV with at minimum 'title' + 'abstract' columns.
    Also accepts 'Title'/'Abstract' (case-insensitive).
    Optional columns: doi, authors, year, journal, keywords.
    """
    import pandas as pd
    fname = Path(path).name
    df    = pd.read_csv(path, encoding="utf-8", on_bad_lines="skip")

    # Case-insensitive column map
    col_map = {c.lower(): c for c in df.columns}

    def _col(name: str) -> str:
        """Return actual column name for a lowercase alias, or ''."""
        return col_map.get(name, "")

    title_col = _col("title")
    if not title_col:
        raise ValueError("CSV file must contain a 'title' column.")

    abstract_col = _col("abstract")
    records = []

    for _, row in df.iterrows():
        r = _empty_record()
        r["source_file"] = fname
        r["title"]    = str(row[title_col]).strip() if title_col else ""
        r["abstract"] = str(row[abstract_col]).strip() if abstract_col else ""
        for field_name in ("doi", "authors", "year", "journal", "keywords"):
            col = _col(field_name)
            if col:
                r[field_name] = str(row[col]).strip()
        r["record_id"] = _make_id(r)
        if r["title"] and r["title"].lower() != "nan":
            records.append(r)

    return records


# ─────────────────────────────────────────────────────────────────────────────
# Public: parse any reference file
# ─────────────────────────────────────────────────────────────────────────────

def parse_references(path: str) -> List[Dict]:
    """
    Auto-detect file type from extension and parse to normalised records.
    Supports .ris, .bib, .csv, .txt (assumed RIS).
    """
    ext = Path(path).suffix.lower()
    if ext in (".ris", ".txt"):
        return _parse_ris(path)
    elif ext in (".bib",):
        return _parse_bib(path)
    elif ext == ".csv":
        return _parse_csv_file(path)
    else:
        # Try RIS as default
        try:
            return _parse_ris(path)
        except Exception:
            raise ValueError(f"Unsupported file format: {ext}")


# ─────────────────────────────────────────────────────────────────────────────
# Deduplication
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class DeduplicationStats:
    total_before: int = 0
    removed_doi:  int = 0
    removed_fuzzy: int = 0
    total_after:  int = 0


def deduplicate(records: List[Dict], fuzzy_threshold: int = 90) -> Tuple[List[Dict], DeduplicationStats]:
    """
    Two-pass deduplication:
      Pass 1 — Exact DOI match (fastest, most reliable)
      Pass 2 — Fuzzy title + author match using thefuzz (catches datasets with
                missing DOIs, slightly different titles from different databases)

    Returns (deduplicated_records, stats).
    """
    stats = DeduplicationStats(total_before=len(records))
    seen_ids: set = set()
    pass1: List[Dict] = []

    # Pass 1: DOI dedup
    for r in records:
        doi = (r.get("doi") or "").strip().lower()
        if doi:
            if doi in seen_ids:
                stats.removed_doi += 1
                continue
            seen_ids.add(doi)
        pass1.append(r)

    # Pass 2: fuzzy title dedup
    try:
        from thefuzz import fuzz
        have_fuzz = True
    except ImportError:
        logger.warning("thefuzz not installed — skipping fuzzy deduplication")
        have_fuzz = False

    if not have_fuzz:
        stats.total_after = len(pass1)
        return pass1, stats

    kept: List[Dict] = []
    seen_titles: List[str] = []

    for r in pass1:
        title = _normalise_title(r.get("title", ""))
        if not title:
            kept.append(r)
            continue
        is_dup = any(
            fuzz.token_sort_ratio(title, t) >= fuzzy_threshold
            for t in seen_titles
        )
        if is_dup:
            stats.removed_fuzzy += 1
        else:
            seen_titles.append(title)
            kept.append(r)

    stats.total_after = len(kept)
    return kept, stats


def _normalise_title(title: str) -> str:
    return re.sub(r'[^\w\s]', '', title.lower()).strip()


def _make_id(r: Dict) -> str:
    src = (r.get("title", "") + r.get("doi", "") + r.get("year", "")).encode()
    return hashlib.md5(src).hexdigest()[:12]


# ─────────────────────────────────────────────────────────────────────────────
# Abstract Screening
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class AbstractScreeningResult:
    record_id:  str
    title:      str
    decision:   str   = "Pending"   # Include / Exclude / Flag / Error
    rationale:  str   = ""
    confidence: str   = ""          # High / Medium / Low
    tokens:     int   = 0
    proc_time:  float = 0.0


DECISION_INCLUDE = "Include"
DECISION_EXCLUDE = "Exclude"
DECISION_FLAG    = "Flag for Human Review"
DECISION_ERROR   = "Error"


class AbstractScreener:
    """
    Screens a list of normalised records using title + abstract alone.
    No PDFs required at this stage.
    """

    SCREENING_PROMPT = (
        "You are an expert academic researcher conducting a systematic review.\n"
        "Your task: decide whether this paper should be INCLUDED, EXCLUDED, or FLAGGED\n"
        "for human review, based solely on the title and abstract.\n\n"
        "REVIEW CRITERIA:\n{criteria}\n\n"
        "Paper:\n"
        "Title: {title}\n"
        "Abstract: {abstract}\n\n"
        "Return ONLY a valid JSON object (no prose, no markdown fences):\n"
        '{{\n'
        '  "decision": "Include | Exclude | Flag for Human Review",\n'
        '  "rationale": "One concise sentence explaining the decision.",\n'
        '  "confidence": "High | Medium | Low"\n'
        '}}'
    )

    def __init__(self, llm_manager, max_retries: int = 3,
                 retry_delay: float = 1.0, rate_limit_delay: float = 0.5):
        self.llm    = llm_manager
        self.max_retries    = max_retries
        self.retry_delay    = retry_delay
        self.rate_limit_delay = rate_limit_delay
        self._stop  = threading.Event()

    def stop(self):
        self._stop.set()

    def screen_all(
        self,
        records:  List[Dict],
        criteria: str,
        callback: Optional[Callable[[AbstractScreeningResult, int, int], None]] = None,
    ) -> List[AbstractScreeningResult]:
        """
        Screen records sequentially (rate-limit safe).
        callback(result, current_index, total) called after each record.
        """
        self._stop.clear()
        results: List[AbstractScreeningResult] = []
        total = len(records)

        for i, rec in enumerate(records):
            if self._stop.is_set():
                break
            res = self._screen_one(rec, criteria)
            results.append(res)
            if callback:
                callback(res, i + 1, total)
            time.sleep(self.rate_limit_delay)

        return results

    def _screen_one(self, rec: Dict, criteria: str) -> AbstractScreeningResult:
        start = time.time()
        title    = (rec.get("title") or "").strip()
        abstract = (rec.get("abstract") or "").strip()
        rid      = rec.get("record_id", "")

        if not title:
            return AbstractScreeningResult(
                record_id=rid, title=title,
                decision=DECISION_ERROR, rationale="No title available",
            )

        prompt = self.SCREENING_PROMPT.format(
            criteria=criteria or "[No criteria defined — include all]",
            title=title,
            abstract=abstract or "Not available",
        )
        messages = [
            {"role": "system", "content": "You screen academic papers for systematic reviews. Return only valid JSON."},
            {"role": "user",   "content": prompt},
        ]

        for attempt in range(self.max_retries):
            if self._stop.is_set():
                return AbstractScreeningResult(record_id=rid, title=title,
                                               decision=DECISION_ERROR,
                                               rationale="Stopped by user")
            try:
                text, tokens = self.llm.chat_completion_with_tokens(
                    messages, temperature=0.05, max_tokens=300
                )
                data = _parse_screening_json(text)
                decision = _normalise_decision(data.get("decision", ""))
                return AbstractScreeningResult(
                    record_id=rid,
                    title=title,
                    decision=decision,
                    rationale=data.get("rationale", ""),
                    confidence=data.get("confidence", ""),
                    tokens=tokens,
                    proc_time=time.time() - start,
                )
            except Exception as exc:
                err = str(exc).lower()
                is_rate = any(x in err for x in ("rate limit", "429", "too many"))
                wait = (2 ** attempt) * (self.retry_delay * (3 if is_rate else 1))
                if attempt < self.max_retries - 1:
                    time.sleep(wait)
                else:
                    return AbstractScreeningResult(
                        record_id=rid, title=title,
                        decision=DECISION_ERROR, rationale=str(exc),
                        proc_time=time.time() - start,
                    )

        return AbstractScreeningResult(record_id=rid, title=title,
                                       decision=DECISION_ERROR,
                                       rationale="Max retries exceeded")


def _parse_screening_json(raw: str) -> Dict:
    cleaned = re.sub(r'```(?:json)?', '', raw, flags=re.IGNORECASE)
    cleaned = re.sub(r'```', '', cleaned).strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    m = re.search(r'\{.*\}', cleaned, re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except json.JSONDecodeError:
            pass
    return {"decision": DECISION_FLAG, "rationale": raw[:300], "confidence": "Low"}


def _normalise_decision(raw: str) -> str:
    raw = (raw or "").strip().lower()
    if "exclude" in raw:
        return DECISION_EXCLUDE
    if "include" in raw and "exclude" not in raw:
        return DECISION_INCLUDE
    if "flag" in raw or "human" in raw:
        return DECISION_FLAG
    return DECISION_FLAG


# ─────────────────────────────────────────────────────────────────────────────
# Export helpers
# ─────────────────────────────────────────────────────────────────────────────

def _records_with_screening(records: List[Dict], results: List[AbstractScreeningResult]) -> List[Dict]:
    """Merge screening decisions back into the record dicts by record_id."""
    res_map = {r.record_id: r for r in results}
    merged = []
    for rec in records:
        r = dict(rec)
        sr = res_map.get(rec["record_id"])
        if sr:
            r["decision"]   = sr.decision
            r["rationale"]  = sr.rationale
            r["confidence"] = sr.confidence
        merged.append(r)
    return merged


EXPORT_COLUMNS = [
    "title", "authors", "year", "journal", "doi",
    "decision", "rationale", "confidence",
    "abstract", "keywords", "source_file", "record_id",
]


def export_records_to_csv(records: List[Dict], results: List[AbstractScreeningResult], path: str):
    merged = _records_with_screening(records, results)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(merged)


def export_records_to_excel(
    records:  List[Dict],
    results:  List[AbstractScreeningResult],
    path:     str,
    stats:    Optional[DeduplicationStats] = None,
):
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(f"Export requires pandas + openpyxl: {e}")

    merged = _records_with_screening(records, results)
    df = pd.DataFrame(merged)

    # Keep only / reorder columns
    cols = [c for c in EXPORT_COLUMNS if c in df.columns]
    df = df[cols]

    df.to_excel(path, index=False, sheet_name="Abstract Screening")

    wb = load_workbook(path)
    ws = wb.active

    FILLS = {
        DECISION_INCLUDE: PatternFill("solid", fgColor="C6EFCE"),
        DECISION_EXCLUDE: PatternFill("solid", fgColor="FFC7CE"),
        DECISION_FLAG:    PatternFill("solid", fgColor="FFEB9C"),
        DECISION_ERROR:   PatternFill("solid", fgColor="DDEBF7"),
    }

    # Style header
    hdr_fill = PatternFill("solid", fgColor="2D6A9F")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # Find decision column index
    dec_col = None
    for i, cell in enumerate(ws[1], start=1):
        if str(cell.value).lower() == "decision":
            dec_col = i
            break

    # Color rows
    for row in ws.iter_rows(min_row=2):
        decision = str(row[dec_col - 1].value) if dec_col else ""
        fill = FILLS.get(decision)
        if fill:
            for cell in row:
                cell.fill = fill

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"

    # PRISMA summary sheet
    if stats or results:
        ws2 = wb.create_sheet("PRISMA Summary")
        rows = [("Stage", "Count")]
        if stats:
            rows += [
                ("Records imported",         stats.total_before),
                ("Duplicates removed (DOI)",  stats.removed_doi),
                ("Duplicates removed (fuzzy)" , stats.removed_fuzzy),
                ("Records after deduplication", stats.total_after),
            ]
        if results:
            from collections import Counter
            dec_counts = Counter(r.decision for r in results)
            rows += [
                ("Included at abstract screening", dec_counts.get(DECISION_INCLUDE, 0)),
                ("Excluded at abstract screening",  dec_counts.get(DECISION_EXCLUDE, 0)),
                ("Flagged for human review",         dec_counts.get(DECISION_FLAG, 0)),
                ("Errors",                           dec_counts.get(DECISION_ERROR, 0)),
            ]
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                cell = ws2.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = Font(bold=True)
        ws2.column_dimensions["A"].width = 42
        ws2.column_dimensions["B"].width = 12

    wb.save(path)
