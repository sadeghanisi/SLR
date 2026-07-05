import csv

from openpyxl import load_workbook

from housing_enhanced import ScreeningDecision, ScreeningResult, SystematicReviewAutomation


def _automation(monkeypatch, tmp_path, *, include_subfolders=False):
    monkeypatch.setattr(SystematicReviewAutomation, "_init_llm", lambda self: object())
    pdf_dir = tmp_path / "pdfs"
    pdf_dir.mkdir()
    auto = SystematicReviewAutomation(
        api_key="test-key",
        pdf_folder=str(pdf_dir),
        output_folder=str(tmp_path / "out"),
        cache_enabled=False,
        parallel_processing=False,
        include_subfolders=include_subfolders,
    )
    return auto, pdf_dir


def _pdf(path):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"%PDF-1.4\n")


def _capture_processed_names(monkeypatch, auto):
    processed = []

    def fake_process_one(pdf_path):
        name = auto._pdf_display_name(pdf_path)
        processed.append(name)
        return ScreeningResult(
            filename=name,
            decision=ScreeningDecision.LIKELY_EXCLUDE.value,
            reasoning="fixture",
            notes="",
        ), None

    monkeypatch.setattr(auto, "_process_one", fake_process_one)
    return processed


def test_direct_folder_pdfs_are_discovered_without_recursive(monkeypatch, tmp_path):
    auto, pdf_dir = _automation(monkeypatch, tmp_path)
    _pdf(pdf_dir / "b.pdf")
    _pdf(pdf_dir / "a.pdf")
    _pdf(pdf_dir / "nested" / "ignored.pdf")
    processed = _capture_processed_names(monkeypatch, auto)

    summary = auto.process_pdfs()

    assert processed == ["a.pdf", "b.pdf"]
    assert summary["statistics"]["total_files"] == 2


def test_subfolder_pdfs_are_included_when_recursive_enabled(monkeypatch, tmp_path):
    auto, pdf_dir = _automation(monkeypatch, tmp_path, include_subfolders=True)
    _pdf(pdf_dir / "root.pdf")
    _pdf(pdf_dir / "nested" / "paper.pdf")
    _pdf(pdf_dir / "other" / "paper.pdf")
    processed = _capture_processed_names(monkeypatch, auto)

    summary = auto.process_pdfs()

    assert processed == ["nested/paper.pdf", "other/paper.pdf", "root.pdf"]
    assert summary["statistics"]["total_files"] == 3


def test_subfolder_pdfs_are_ignored_when_recursive_disabled(monkeypatch, tmp_path):
    auto, pdf_dir = _automation(monkeypatch, tmp_path)
    _pdf(pdf_dir / "nested" / "ignored.pdf")
    processed = _capture_processed_names(monkeypatch, auto)

    summary = auto.process_pdfs()

    assert processed == []
    assert summary["statistics"]["total_files"] == 0


def test_process_one_uses_relative_name_for_nested_pdf(monkeypatch, tmp_path):
    auto, pdf_dir = _automation(monkeypatch, tmp_path, include_subfolders=True)
    pdf_path = pdf_dir / "nested" / "paper.pdf"
    _pdf(pdf_path)
    text = "Nested paper text."

    def fake_extract(path):
        auto._record_text_extraction_metadata(
            path,
            method="fake",
            status="ok",
            text=text,
            page_count=1,
        )
        return text, True

    monkeypatch.setattr(auto, "extract_text_from_pdf", fake_extract)
    monkeypatch.setattr(
        auto,
        "screen_article",
        lambda paper_text, filename, stage="Full-text": ScreeningResult(
            filename=filename,
            decision=ScreeningDecision.LIKELY_EXCLUDE.value,
            reasoning="fixture",
            notes="",
            stage=stage,
            text_length=len(paper_text),
        ),
    )

    screening, extraction = auto._process_one(pdf_path)

    assert screening.filename == "nested/paper.pdf"
    assert extraction is None
    assert auto.stats["current_file"] == "nested/paper.pdf"
    assert auto.get_paper_text_metadata("nested/paper.pdf")["original_filename"] == "nested/paper.pdf"


def test_same_basename_nested_pdfs_remain_distinct_in_csv(monkeypatch, tmp_path):
    auto, pdf_dir = _automation(monkeypatch, tmp_path, include_subfolders=True)
    _pdf(pdf_dir / "alpha" / "paper.pdf")
    _pdf(pdf_dir / "beta" / "paper.pdf")
    _capture_processed_names(monkeypatch, auto)

    summary = auto.process_pdfs()

    with open(auto.screening_csv, newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))

    assert summary["statistics"]["total_files"] == 2
    assert [row["filename"] for row in rows] == ["alpha/paper.pdf", "beta/paper.pdf"]


def test_same_basename_nested_pdfs_remain_distinct_in_excel_and_summary(monkeypatch, tmp_path):
    auto, pdf_dir = _automation(monkeypatch, tmp_path, include_subfolders=True)
    _pdf(pdf_dir / "alpha" / "paper.pdf")
    _pdf(pdf_dir / "beta" / "paper.pdf")
    _capture_processed_names(monkeypatch, auto)

    summary = auto.process_pdfs()

    workbook = load_workbook(auto.screening_excel)
    worksheet = workbook["Screening Results"]
    filenames = [
        worksheet.cell(row=row_index, column=1).value
        for row_index in range(2, worksheet.max_row + 1)
    ]

    assert filenames == ["alpha/paper.pdf", "beta/paper.pdf"]
    assert summary["screening_excel"] == str(auto.screening_excel)
    assert summary["screened_count"] == 2
    assert auto.summary_report.exists()
