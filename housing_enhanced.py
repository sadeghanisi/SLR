"""
Universal Systematic / Scoping Review Automation — Backend
Domain-agnostic for any academic research field.
housing_enhanced.py is kept as the module name for import compatibility.
"""

__version__ = "3.1.0"

import os
import csv
import re
import time
import threading
import traceback
import json
import unicodedata
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import hashlib
from llm_interface import LLMManager
from dataclasses import dataclass, asdict, field
from enum import Enum

CACHE_SCHEMA_VERSION = "2"


def _canonical_json(data: Any) -> str:
    """Stable JSON representation for reproducible hashes."""
    return json.dumps(
        data,
        sort_keys=True,
        ensure_ascii=True,
        separators=(",", ":"),
        default=str,
    )


def _hash_utf8(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _normalize_text_for_cache(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", text or "")
    normalized = normalized.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"[ \t]+", " ", normalized)
    normalized = re.sub(r"\n{3,}", "\n\n", normalized)
    return normalized.strip()


def _sha256_text(text: str) -> str:
    return _hash_utf8(_normalize_text_for_cache(text))


def _sha256_json(data: Any) -> str:
    return _hash_utf8(_canonical_json(data))


def _without_secrets(data: Dict[str, Any]) -> Dict[str, Any]:
    secret_terms = ("key", "token", "secret", "password", "credential")
    return {
        key: value
        for key, value in data.items()
        if not any(term in key.lower() for term in secret_terms)
    }


# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────

def setup_logging(log_file: str = 'slr_automation.log') -> logging.Logger:
    """Setup enhanced logging configuration"""
    logger = logging.getLogger('SLR_Automation')
    logger.setLevel(logging.INFO)
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    formatter = logging.Formatter(
        '%(asctime)s [%(levelname)s] %(funcName)s: %(message)s'
    )
    try:
        fh = logging.FileHandler(log_file, encoding='utf-8')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
    except Exception:
        pass
    ch = logging.StreamHandler()
    ch.setFormatter(formatter)
    logger.addHandler(ch)
    return logger


# ─────────────────────────────────────────────────────────────────────────────
# Enumerations & Data Classes
# ─────────────────────────────────────────────────────────────────────────────

class ScreeningDecision(Enum):
    LIKELY_INCLUDE  = "Likely Include"
    LIKELY_EXCLUDE  = "Likely Exclude"
    FLAG_FOR_REVIEW = "Flag for Review"
    FLAG_FOR_HUMAN  = "Flag for Human Review"
    ERROR           = "Error"


@dataclass
class ScreeningResult:
    filename:        str
    decision:        str
    reasoning:       str
    notes:           str
    stage:           str   = "Full-text"   # "Title/Abstract" or "Full-text"
    processing_time: float = 0.0
    text_length:     int   = 0
    api_tokens_used: int   = 0


@dataclass
class ExtractionResult:
    """
    Dynamic extraction result — fields dict holds user-defined columns.
    No hardcoded domain-specific fields.
    """
    filename:        str
    fields:          Dict[str, Any] = field(default_factory=dict)
    processing_time: float = 0.0
    api_tokens_used: int   = 0


@dataclass
class TextCacheMetadata:
    original_filename:    str
    text_hash:            str
    text_cache_path:      str
    extraction_method:    str = "unknown"
    extraction_status:    str = "unknown"
    extracted_char_count: int = 0
    page_count:           Optional[int] = None


# ─────────────────────────────────────────────────────────────────────────────
# Robust JSON parser
# ─────────────────────────────────────────────────────────────────────────────

def _parse_json_response(raw: str) -> Dict:
    """
    Extract JSON from an LLM response even when wrapped in markdown fences,
    mixed with prose, or containing minor syntax errors.
    """
    if not raw:
        raise ValueError("Empty LLM response")

    # Strip ```json ... ``` or ``` ... ```
    cleaned = re.sub(r'```(?:json)?\s*', '', raw, flags=re.IGNORECASE)
    cleaned = re.sub(r'```', '', cleaned).strip()

    # Direct parse
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # Extract first {...} block
    m = re.search(r'\{.*\}', cleaned, re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except json.JSONDecodeError:
            pass

    # Fallback so processing doesn't crash
    return {
        "decision": ScreeningDecision.FLAG_FOR_HUMAN.value,
        "reasoning": "Could not parse LLM response as JSON.",
        "notes": f"Raw response (first 600 chars): {raw[:600]}"
    }


# ─────────────────────────────────────────────────────────────────────────────
# Main Automation Class
# ─────────────────────────────────────────────────────────────────────────────

class SystematicReviewAutomation:

    def __init__(
        self,
        api_key:              str,
        pdf_folder:           str,
        output_folder:        str  = "output",
        cache_enabled:        bool = True,
        parallel_processing:  bool = True,
        max_workers:          int  = 3,
        rate_limit_delay:     float = 1.0,
        screening_prompt:     str  = None,
        extraction_prompt:    str  = None,
        extraction_fields:    List[str] = None,
        llm_provider:         str  = "OpenAI",
        llm_model:            str  = None,
        two_stage_screening:  bool = False,
        stop_event:           threading.Event = None,
        advanced_config:      dict = None,
        **llm_kwargs,
    ):
        self.api_key             = api_key
        self.pdf_folder          = Path(pdf_folder)
        self.output_folder       = Path(output_folder)
        self.cache_enabled       = cache_enabled
        self.parallel_processing = parallel_processing
        self.max_workers         = max_workers
        self.rate_limit_delay    = rate_limit_delay
        self.llm_provider        = llm_provider
        self.llm_model           = llm_model or LLMManager.get_default_models().get(llm_provider, "gpt-5.5")
        self.llm_kwargs          = llm_kwargs
        self.two_stage_screening = two_stage_screening
        self.stop_event          = stop_event or threading.Event()

        adv = advanced_config or {}
        self.max_text_chars             = adv.get('max_text_chars', 100000)
        self.max_retries                = adv.get('max_retries', 3)
        self.retry_delay_base           = adv.get('retry_delay', 0.5)
        self.intermediate_save_interval = adv.get('intermediate_save_interval', 5)
        self.enable_smart_truncation    = adv.get('enable_smart_truncation', True)
        self.preserve_sections          = adv.get('preserve_sections', [
            'abstract', 'introduction', 'method', 'result', 'discussion', 'conclusion'
        ])
        self.strip_references           = adv.get('strip_references', True)
        self._advanced_config_for_cache = {
            'max_text_chars': self.max_text_chars,
            'max_retries': self.max_retries,
            'retry_delay': self.retry_delay_base,
            'intermediate_save_interval': self.intermediate_save_interval,
            'enable_smart_truncation': self.enable_smart_truncation,
            'preserve_sections': self.preserve_sections,
            'strip_references': self.strip_references,
        }

        # Extraction fields define the CSV/Excel columns for extracted data
        self.extraction_fields = extraction_fields or [
            "title", "authors", "publication_year", "study_design",
            "study_objectives", "methodology", "population", "sample_size",
            "main_findings", "conclusions", "limitations", "keywords"
        ]

        # Prompts
        self.screening_prompt  = screening_prompt  or self._default_screening_prompt()
        self.extraction_prompt = extraction_prompt or self._default_extraction_prompt()

        # Create output dir early so logger can write there
        self.output_folder.mkdir(parents=True, exist_ok=True)
        self.logger = setup_logging(str(self.output_folder / 'slr_automation.log'))

        # LLM
        self.llm_manager = self._init_llm()

        # Cache
        self.cache_folder = self.output_folder / 'cache'
        if self.cache_enabled:
            self.cache_folder.mkdir(exist_ok=True)
        self.text_cache_folder = self.output_folder / 'text_cache'
        self.text_cache_folder.mkdir(exist_ok=True)

        # Timestamped output files
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.screening_csv    = self.output_folder / f"screening_{ts}.csv"
        self.extraction_csv   = self.output_folder / f"extraction_{ts}.csv"
        self.screening_excel  = self.output_folder / f"screening_{ts}.xlsx"
        self.extraction_excel = self.output_folder / f"extraction_{ts}.xlsx"
        self.summary_report   = self.output_folder / f"summary_{ts}.txt"
        self.audit_ledger     = self.output_folder / f"audit_{ts}.jsonl"

        # Results
        self.screening_results:  List[ScreeningResult]  = []
        self.extraction_results: List[ExtractionResult] = []
        self._paper_texts:       Dict[str, str]         = {}  # legacy compatibility; not populated
        self._paper_text_metadata: Dict[str, Dict[str, Any]] = {}
        self._last_text_extraction_metadata: Dict[str, Dict[str, Any]] = {}

        # Thread-safe stats
        self._lock = threading.Lock()
        self._audit_lock = threading.Lock()
        self.stats = {
            'total_files':         0,
            'processed_files':     0,
            'failed_files':        0,
            'likely_include':      0,
            'likely_exclude':      0,
            'flag_for_review':     0,
            'flag_for_human_review': 0,
            'error':               0,
            'total_processing_time': 0.0,
            'total_api_tokens':    0,
            'current_file':        '',
        }

    # ── LLM init ────────────────────────────────────────────────────────────

    def _init_llm(self) -> LLMManager:
        try:
            mgr = LLMManager(
                provider_name=self.llm_provider,
                api_key=self.api_key,
                model=self.llm_model,
                **self.llm_kwargs,
            )
            self.logger.info(f"LLM ready: {self.llm_provider} / {self.llm_model}")
            return mgr
        except Exception as e:
            self.logger.error(f"LLM init failed: {e}")
            raise

    # ── Default generic prompts ──────────────────────────────────────────────

    def _default_screening_prompt(self) -> str:
        return (
            "You are an expert AI assistant supporting a systematic or scoping review.\n"
            "Read the research paper below and decide if it meets the review's criteria.\n\n"
            "INCLUSION CRITERIA:\n"
            "  [Edit in the Customize Criteria dialog — e.g. study type, population, intervention, date range]\n\n"
            "EXCLUSION CRITERIA:\n"
            "  [Edit in the Customize Criteria dialog — e.g. animal studies, non-English, editorials]\n\n"
            "Paper text:\n{text}\n\n"
            "Return ONLY a valid JSON object (no prose, no markdown):\n"
            '{{\n'
            '    "decision": "Likely Include | Likely Exclude | Flag for Review | Flag for Human Review",\n'
            '    "reasoning": "Step-by-step reasoning against each criterion",\n'
            '    "notes": "Quality concerns, missing info, or caveats"\n'
            '}}'
        )

    def _default_extraction_prompt(self) -> str:
        fields_lines = "\n".join(
            f'    "{f}": "extracted value or Not reported",'
            for f in self.extraction_fields
        )
        return (
            "You are an expert data extractor for academic systematic/scoping reviews.\n"
            "Extract the information below from the research paper.\n"
            "Use 'Not reported' when a field cannot be determined from the text.\n\n"
            "Paper text:\n{text}\n\n"
            "Return ONLY a valid JSON object (no prose, no markdown):\n"
            "{{\n"
            f"{fields_lines}\n"
            "}}"
        )

    # ── Cache helpers ────────────────────────────────────────────────────────

    def _provider_cache_metadata(self) -> Dict[str, Any]:
        return {
            'provider': self.llm_provider,
            'model': self.llm_model,
            'base_url': _without_secrets(self.llm_kwargs).get('base_url'),
        }

    def _extraction_prompt_fingerprint_source(self) -> Dict[str, Any]:
        return {
            'prompt': self.extraction_prompt,
            'quote_then_answer': True,
            'fields': self.extraction_fields,
        }

    def _cache_key_context(
        self,
        *,
        kind: str,
        text: str,
        prompt: Any,
        stage: str = None,
    ) -> Dict[str, Any]:
        context = {
            'cache_schema_version': CACHE_SCHEMA_VERSION,
            'app_version': __version__,
            'pipeline_version': __version__,
            'kind': kind,
            'stage': stage,
            'text_hash': _sha256_text(text),
            'provider_profile': self._provider_cache_metadata(),
            'model': self.llm_model,
            'prompt_hash': _sha256_json(prompt),
            'extraction_fields_hash': _sha256_json(self.extraction_fields),
            'advanced_config_hash': _sha256_json(self._advanced_config_for_cache),
        }
        context['cache_key'] = _sha256_json(context)
        return context

    def _write_audit_event(
        self,
        *,
        kind: str,
        filename: str,
        cache_context: Dict[str, Any],
        cache_hit: bool,
        stage: str = None,
        tokens: int = 0,
        decision: str = None,
        status: str = "ok",
        parse_status: str = "ok",
        retry_count: int = 0,
        error: str = None,
    ):
        provider_profile = cache_context.get('provider_profile') or {}
        text_cache_metadata = self.get_paper_text_metadata(filename)
        event = {
            'timestamp': datetime.now().isoformat(timespec='seconds'),
            'schema_version': CACHE_SCHEMA_VERSION,
            'app_version': __version__,
            'pipeline_version': __version__,
            'kind': kind,
            'filename': filename,
            'stage': stage,
            'cache_hit': cache_hit,
            'cache_key': cache_context.get('cache_key'),
            'text_hash': cache_context.get('text_hash'),
            'provider': provider_profile.get('provider'),
            'provider_profile': cache_context.get('provider_profile'),
            'model': cache_context.get('model'),
            'prompt_hash': cache_context.get('prompt_hash'),
            'extraction_fields_hash': cache_context.get('extraction_fields_hash'),
            'advanced_config_hash': cache_context.get('advanced_config_hash'),
            'api_tokens_used': tokens,
            'decision': decision,
            'status': status,
            'parse_status': parse_status,
            'retry_count': retry_count,
        }
        if text_cache_metadata:
            event['text_cache'] = text_cache_metadata
        if error:
            event['error'] = error
        try:
            with self._audit_lock:
                with open(self.audit_ledger, 'a', encoding='utf-8') as f:
                    f.write(_canonical_json(event) + '\n')
        except Exception as e:
            self.logger.warning(f"Audit ledger write: {e}")

    def _cache_metadata_matches(
        self,
        data: Dict[str, Any],
        expected_context: Optional[Dict[str, Any]],
        kind: str,
    ) -> bool:
        if expected_context is None:
            return True
        metadata = data.get('cache_metadata')
        if not isinstance(metadata, dict):
            return False
        required = (
            'cache_schema_version',
            'app_version',
            'pipeline_version',
            'kind',
            'stage',
            'text_hash',
            'provider_profile',
            'model',
            'prompt_hash',
            'extraction_fields_hash',
            'advanced_config_hash',
            'cache_key',
        )
        if any(field not in metadata for field in required):
            return False
        if metadata.get('cache_schema_version') != CACHE_SCHEMA_VERSION:
            return False
        if metadata.get('kind') != kind:
            return False
        return all(metadata.get(field) == value for field, value in expected_context.items())

    def _cache_load(
        self,
        key: str,
        kind: str,
        expected_context: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict]:
        if not self.cache_enabled:
            return None
        # Runtime cache loading is JSON-only. Legacy pickle caches are unsafe to
        # load automatically; migrate them manually from a trusted environment.
        jp = self.cache_folder / f"{kind}_{key}.json"
        if jp.exists():
            try:
                with open(jp, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if expected_context is None:
                    return data if isinstance(data, dict) else None
                if isinstance(data, dict) and self._cache_metadata_matches(data, expected_context, kind):
                    return data
            except Exception:
                pass
        return None

    def _cache_save(self, key: str, kind: str, data: Dict):
        if not self.cache_enabled:
            return
        try:
            with open(self.cache_folder / f"{kind}_{key}.json", 'w', encoding='utf-8') as f:
                json.dump(data, f)
        except Exception as e:
            self.logger.warning(f"Cache write: {e}")

    # ── Disk-backed extracted text cache ────────────────────────────────────

    def _record_text_extraction_metadata(
        self,
        pdf_path: Path,
        *,
        method: str,
        status: str,
        text: str = "",
        page_count: Optional[int] = None,
    ) -> Dict[str, Any]:
        metadata = {
            'original_filename': pdf_path.name,
            'extraction_method': method,
            'extraction_status': status,
            'extracted_char_count': len(text or ""),
            'page_count': page_count,
        }
        self._last_text_extraction_metadata[pdf_path.name] = metadata
        return metadata

    def _store_paper_text_cache(
        self,
        filename: str,
        text: str,
        extraction_metadata: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        extraction_metadata = extraction_metadata or {}
        text_hash = _hash_utf8(text or "")
        cache_path = self.text_cache_folder / f"{text_hash}.txt"
        try:
            self.text_cache_folder.mkdir(parents=True, exist_ok=True)
            if not cache_path.exists():
                cache_path.write_text(text, encoding='utf-8')
        except Exception as e:
            self.logger.warning(f"Text cache write failed ({filename}): {e}")

        metadata = TextCacheMetadata(
            original_filename=filename,
            text_hash=text_hash,
            text_cache_path=str(cache_path),
            extraction_method=extraction_metadata.get('extraction_method', 'unknown'),
            extraction_status=extraction_metadata.get('extraction_status', 'ok'),
            extracted_char_count=len(text),
            page_count=extraction_metadata.get('page_count'),
        )
        self._paper_text_metadata[filename] = asdict(metadata)
        metadata_path = self.text_cache_folder / f"{text_hash}.json"
        try:
            metadata_path.write_text(
                _canonical_json(self._paper_text_metadata[filename]),
                encoding='utf-8',
            )
        except Exception as e:
            self.logger.warning(f"Text cache metadata write failed ({filename}): {e}")
        return dict(self._paper_text_metadata[filename])

    def _record_paper_text_failure(
        self,
        filename: str,
        extraction_metadata: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        extraction_metadata = extraction_metadata or {}
        metadata = TextCacheMetadata(
            original_filename=filename,
            text_hash="",
            text_cache_path="",
            extraction_method=extraction_metadata.get('extraction_method', 'unknown'),
            extraction_status=extraction_metadata.get('extraction_status', 'error'),
            extracted_char_count=0,
            page_count=extraction_metadata.get('page_count'),
        )
        self._paper_text_metadata[filename] = asdict(metadata)
        return dict(self._paper_text_metadata[filename])

    def get_paper_text_metadata(self, filename: str) -> Dict[str, Any]:
        return dict(self._paper_text_metadata.get(filename, {}))

    def get_paper_text(self, filename: str) -> str:
        metadata = self._paper_text_metadata.get(filename)
        if not metadata:
            return self._paper_texts.get(filename, "")

        cache_path_value = metadata.get('text_cache_path')
        if not cache_path_value:
            return "(Paper text not available; extraction did not produce cached text.)"

        try:
            cache_path = Path(cache_path_value).resolve()
            cache_path.relative_to(self.text_cache_folder.resolve())
            return cache_path.read_text(encoding='utf-8')
        except FileNotFoundError:
            return "(Paper text cache missing; rerun processing to rebuild it.)"
        except Exception as e:
            self.logger.warning(f"Text cache read failed ({filename}): {e}")
            return "(Paper text unavailable; text cache could not be read.)"

    # ── Text utilities ───────────────────────────────────────────────────────

    def _smart_truncate(self, text: str, max_chars: int) -> str:
        if len(text) <= max_chars:
            return text
        if not self.enable_smart_truncation:
            return text[:max_chars]
        paragraphs = text.split('\n\n')
        pri, rest = [], []
        kw = self.preserve_sections
        for p in paragraphs:
            (pri if any(k in p.lower() for k in kw) else rest).append(p)
        result = '\n\n'.join(pri)
        for p in rest:
            if len(result) + len(p) + 2 <= max_chars:
                result += '\n\n' + p
        return result[:max_chars]

    # ── LLM call with retry ──────────────────────────────────────────────────

    def _llm_call(self, messages: List[Dict], **kwargs) -> Tuple[str, int]:
        """Returns (response_text, tokens_used)."""
        for attempt in range(self.max_retries):
            if self.stop_event.is_set():
                raise InterruptedError("Stopped by user")
            try:
                return self.llm_manager.chat_completion_with_tokens(messages, **kwargs)
            except InterruptedError:
                raise
            except Exception as e:
                err = str(e).lower()
                is_rate = any(x in err for x in ("rate limit", "too many requests", "429"))
                is_temp = any(x in err for x in ("timeout", "server error", "502", "503", "504"))
                if attempt == self.max_retries - 1:
                    raise
                wait = (2 ** attempt) * (self.rate_limit_delay if is_rate else self.retry_delay_base)
                self.logger.warning(f"Retry {attempt+1}/{self.max_retries} in {wait:.1f}s: {e}")
                time.sleep(wait)
        raise RuntimeError("Max retries exceeded")

    # ── PDF extraction ───────────────────────────────────────────────────────

    def extract_text_from_pdf(self, pdf_path: Path) -> Tuple[str, bool]:
        """
        Convert PDF to Markdown text using PyMuPDF4LLM (best quality: preserves
        reading order, multi-column layouts, and tables).
        Falls back to pdfplumber → PyPDF2 if PyMuPDF4LLM is unavailable.
        """
        # Primary: PyMuPDF4LLM — academic-grade PDF → Markdown conversion
        try:
            import pymupdf4llm
            md_text = pymupdf4llm.to_markdown(str(pdf_path))
            if md_text.strip():
                if self.strip_references:
                    md_text = self._strip_references_from_md(md_text)
                self._record_text_extraction_metadata(
                    pdf_path,
                    method='pymupdf4llm',
                    status='ok',
                    text=md_text,
                )
                self.logger.info(f"PyMuPDF4LLM: {len(md_text):,} chars from {pdf_path.name}")
                return md_text, True
        except ImportError:
            self.logger.warning("pymupdf4llm not installed — falling back to pdfplumber")
        except Exception as e:
            self.logger.warning(f"PyMuPDF4LLM failed ({pdf_path.name}): {e}")

        # Fallback 1: pdfplumber
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                page_count = len(pdf.pages)
                text = "\n".join((p.extract_text() or "") for p in pdf.pages)
            if text.strip():
                if self.strip_references:
                    text = self._strip_references_from_md(text)
                self._record_text_extraction_metadata(
                    pdf_path,
                    method='pdfplumber',
                    status='ok',
                    text=text,
                    page_count=page_count,
                )
                self.logger.info(f"pdfplumber: {len(text):,} chars from {pdf_path.name}")
                return text, True
        except ImportError:
            pass
        except Exception as e:
            self.logger.warning(f"pdfplumber failed ({pdf_path.name}): {e}")

        # Fallback 2: PyPDF2
        try:
            from PyPDF2 import PdfReader
            with open(pdf_path, 'rb') as fh:
                reader = PdfReader(fh)
                if reader.is_encrypted:
                    self._record_text_extraction_metadata(
                        pdf_path,
                        method='pypdf2',
                        status='encrypted',
                        page_count=len(reader.pages),
                    )
                    return "Error: PDF is encrypted", False
                page_count = len(reader.pages)
                page_texts = []
                for i, page in enumerate(reader.pages):
                    try:
                        page_texts.append(page.extract_text() or "")
                    except Exception as pe:
                        self.logger.warning(f"Page {i+1} skip in {pdf_path.name}: {pe}")
                text = "\n".join(page_texts)
            if text.strip():
                self._record_text_extraction_metadata(
                    pdf_path,
                    method='pypdf2',
                    status='ok',
                    text=text,
                    page_count=page_count,
                )
                self.logger.info(f"PyPDF2: {len(text):,} chars from {pdf_path.name}")
                return text, True
            self._record_text_extraction_metadata(
                pdf_path,
                method='pypdf2',
                status='no_extractable_text',
                page_count=page_count,
            )
            return "Error: No extractable text (possibly a scanned/image-only PDF)", False
        except Exception as e:
            self._record_text_extraction_metadata(
                pdf_path,
                method='unknown',
                status='error',
            )
            return f"Error reading PDF: {e}", False

    def _strip_references_from_md(self, text: str) -> str:
        """
        Heuristically remove the bibliography / references section to reduce
        token usage. Matches Markdown-style headers like '## References'.
        """
        match = re.search(
            r'\n#{1,3}\s*(References|Bibliography|Works Cited|Literature Cited)\s*\n',
            text, flags=re.IGNORECASE
        )
        if match:
            return text[:match.start()].rstrip()
        return text

    def _generate_dynamic_schema(self):
        """
        Build a dynamic Pydantic model implementing the Quote-Then-Answer
        anti-hallucination pattern.

        For each user-defined extraction field (e.g. 'sample_size'), two model
        fields are created:
          • sample_size_quote — verbatim sentence copied from the paper text
          • sample_size       — final extracted value derived ONLY from that quote

        Returns the model class, or None if pydantic is not installed.
        """
        try:
            from pydantic import create_model, Field as PField
        except ImportError:
            return None

        fields_def: Dict[str, Any] = {}
        for field_name in self.extraction_fields:
            safe = field_name.lower().replace(" ", "_").replace("-", "_")
            fields_def[f"{safe}_quote"] = (
                str,
                PField(
                    default="Not found",
                    description=(
                        f"Exact verbatim sentence(s) from the paper text that contain "
                        f"information about '{field_name}'. Copy word-for-word."
                    ),
                ),
            )
            fields_def[safe] = (
                str,
                PField(
                    default="Not reported",
                    description=(
                        f"Extracted value for '{field_name}'. "
                        f"Derive this ONLY from the quote above. "
                        f"Use 'Not reported' if no relevant quote was found."
                    ),
                ),
            )
        return create_model("DynamicExtractionSchema", **fields_def)

    # ── Screening ────────────────────────────────────────────────────────────

    def screen_article(self, text: str, filename: str, stage: str = "Full-text") -> ScreeningResult:
        start = time.time()
        cache_context = self._cache_key_context(
            kind='screening',
            text=text,
            prompt={'screening_prompt': self.screening_prompt, 'stage': stage},
            stage=stage,
        )
        cache_key = cache_context['cache_key']
        cached = self._cache_load(cache_key, 'screening', expected_context=cache_context)
        if cached:
            cached.update({'filename': filename, 'processing_time': time.time() - start})
            # Only pass fields that exist in the dataclass
            valid = {k: cached[k] for k in ScreeningResult.__dataclass_fields__ if k in cached}
            result = ScreeningResult(**valid)
            self._write_audit_event(
                kind='screening',
                filename=filename,
                stage=stage,
                cache_context=cache_context,
                cache_hit=True,
                tokens=0,
                decision=result.decision,
            )
            return result

        try:
            snippet = (
                text[:3000]  # title/abstract only for stage-1
                if stage == "Title/Abstract"
                else self._smart_truncate(text, self.max_text_chars)
            )
            messages = [
                {"role": "system", "content": "You are a systematic review assistant. Only return valid JSON."},
                {"role": "user",   "content": self.screening_prompt.format(text=snippet)},
            ]
            raw, tokens = self._llm_call(messages, temperature=0.05, max_tokens=1500)
            data = _parse_json_response(raw)

            result = ScreeningResult(
                filename=filename,
                decision=data.get("decision", ScreeningDecision.FLAG_FOR_HUMAN.value),
                reasoning=data.get("reasoning", ""),
                notes=data.get("notes", ""),
                stage=stage,
                processing_time=time.time() - start,
                text_length=len(text),
                api_tokens_used=tokens,
            )

            cache_data = asdict(result)
            cache_data.pop('filename', None)
            cache_data.pop('processing_time', None)
            cache_data['cache_metadata'] = cache_context
            self._cache_save(cache_key, 'screening', cache_data)

            with self._lock:
                self.stats['total_api_tokens'] += tokens
                dk = result.decision.lower().replace(' ', '_')
                if dk in self.stats:
                    self.stats[dk] += 1

            self._write_audit_event(
                kind='screening',
                filename=filename,
                stage=stage,
                cache_context=cache_context,
                cache_hit=False,
                tokens=tokens,
                decision=result.decision,
            )
            return result

        except InterruptedError:
            raise
        except Exception as e:
            self.logger.error(f"Screening error {filename}: {e}\n{traceback.format_exc()}")
            self._write_audit_event(
                kind='screening',
                filename=filename,
                stage=stage,
                cache_context=cache_context,
                cache_hit=False,
                status='error',
                parse_status='error',
                error=str(e),
            )
            return ScreeningResult(
                filename=filename,
                decision=ScreeningDecision.ERROR.value,
                reasoning=str(e),
                notes="Screening failed",
                stage=stage,
                processing_time=time.time() - start,
                text_length=len(text),
            )

    # ── Extraction ───────────────────────────────────────────────────────────

    def extract_data(self, text: str, filename: str) -> ExtractionResult:
        start = time.time()
        cache_context = self._cache_key_context(
            kind='extraction',
            text=text,
            prompt=self._extraction_prompt_fingerprint_source(),
        )
        cache_key = cache_context['cache_key']
        cached = self._cache_load(cache_key, 'extraction', expected_context=cache_context)
        if cached:
            result = ExtractionResult(
                filename=filename,
                fields=cached.get('fields', {}),
                processing_time=time.time() - start,
                api_tokens_used=cached.get('api_tokens_used', 0),
            )
            self._write_audit_event(
                kind='extraction',
                filename=filename,
                cache_context=cache_context,
                cache_hit=True,
                tokens=0,
            )
            return result

        try:
            chunk = self._smart_truncate(text, self.max_text_chars)

            # Build Quote-Then-Answer prompt — forces the model to ground each
            # extracted value in a verbatim sentence from the paper.
            qa_lines = ""
            for f in self.extraction_fields:
                safe = f.lower().replace(" ", "_").replace("-", "_")
                qa_lines += (
                    f'    "{safe}_quote": "verbatim sentence(s) from the text about {f}",\n'
                    f'    "{safe}": "extracted value for {f}, or Not reported",\n'
                )
            qa_prompt = (
                "You are an expert data extractor for academic systematic/scoping reviews.\n"
                "Apply the Quote-Then-Answer method for every field:\n"
                "  1. Copy a verbatim sentence from the paper as the field_quote.\n"
                "  2. Write the extracted value based ONLY on that quote.\n"
                "  3. If no relevant text exists, set both field_quote and field to 'Not reported'.\n\n"
                f"Paper text:\n{chunk}\n\n"
                "Return ONLY valid JSON (no prose, no markdown fences):\n"
                "{{\n"
                f"{qa_lines}"
                "}}"
            )
            messages = [
                {"role": "system", "content": "You extract structured data from research papers. Return only valid JSON."},
                {"role": "user",   "content": qa_prompt},
            ]

            # Attempt structured output via instructor (enforces schema + auto-retries)
            schema = self._generate_dynamic_schema()
            data, tokens = None, 0
            if schema is not None:
                try:
                    structured, tokens = self.llm_manager.chat_completion_structured(
                        messages, schema
                    )
                    if structured is not None:
                        data = structured.model_dump()
                except Exception as se:
                    self.logger.warning(
                        f"Structured output failed ({filename}), falling back to JSON parse: {se}"
                    )

            # Fallback: plain LLM call + regex JSON extraction
            if data is None:
                raw, tokens = self._llm_call(messages, temperature=0.05, max_tokens=4000)
                data = _parse_json_response(raw)

            result = ExtractionResult(
                filename=filename,
                fields=data,
                processing_time=time.time() - start,
                api_tokens_used=tokens,
            )

            self._cache_save(cache_key, 'extraction', {
                'fields': data,
                'api_tokens_used': tokens,
                'cache_metadata': cache_context,
            })
            with self._lock:
                self.stats['total_api_tokens'] += tokens
            self._write_audit_event(
                kind='extraction',
                filename=filename,
                cache_context=cache_context,
                cache_hit=False,
                tokens=tokens,
            )
            return result

        except InterruptedError:
            raise
        except Exception as e:
            self.logger.error(f"Extraction error {filename}: {e}\n{traceback.format_exc()}")
            self._write_audit_event(
                kind='extraction',
                filename=filename,
                cache_context=cache_context,
                cache_hit=False,
                status='error',
                parse_status='error',
                error=str(e),
            )
            return ExtractionResult(
                filename=filename,
                fields={f: "Error" for f in self.extraction_fields},
                processing_time=time.time() - start,
            )

    # ── Single-PDF dispatcher ────────────────────────────────────────────────

    def _process_one(self, pdf_file: Path) -> Tuple[Optional[ScreeningResult], Optional[ExtractionResult]]:
        if self.stop_event.is_set():
            return None, None

        with self._lock:
            self.stats['current_file'] = pdf_file.name

        self.logger.info(f"Processing: {pdf_file.name}")
        text, ok = self.extract_text_from_pdf(pdf_file)
        extraction_metadata = self._last_text_extraction_metadata.get(pdf_file.name, {})

        if ok:
            self._store_paper_text_cache(pdf_file.name, text, extraction_metadata)

        if not ok:
            self._record_paper_text_failure(pdf_file.name, extraction_metadata)
            return ScreeningResult(
                filename=pdf_file.name,
                decision="Error",
                reasoning=text,
                notes="PDF extraction failed",
            ), None

        # Optional stage-1: title/abstract quick screen
        if self.two_stage_screening:
            s1 = self.screen_article(text, pdf_file.name, stage="Title/Abstract")
            if s1.decision in (ScreeningDecision.LIKELY_EXCLUDE.value, "Error"):
                return s1, None

        # Full screening
        screening = self.screen_article(text, pdf_file.name, stage="Full-text")

        # Data extraction only for included papers
        extraction = None
        if screening.decision == ScreeningDecision.LIKELY_INCLUDE.value and not self.stop_event.is_set():
            extraction = self.extract_data(text, pdf_file.name)

        time.sleep(self.rate_limit_delay)
        return screening, extraction

    # ── Batch runner ─────────────────────────────────────────────────────────

    def process_pdfs(self) -> Dict:
        pdf_files = list(self.pdf_folder.glob("*.pdf"))
        with self._lock:
            self.stats['total_files'] = len(pdf_files)

        if not pdf_files:
            self.logger.warning(f"No PDFs in {self.pdf_folder}")
            return self._generate_summary()

        self.logger.info(
            f"Batch: {len(pdf_files)} PDFs | "
            f"parallel={self.parallel_processing} workers={self.max_workers} | "
            f"two_stage={self.two_stage_screening}"
        )
        t0 = time.time()

        def record(sr, er):
            if sr:
                self.screening_results.append(sr)
                with self._lock:
                    self.stats['processed_files'] += 1
                    if sr.decision == "Error":
                        self.stats['failed_files'] += 1
            if er:
                self.extraction_results.append(er)
            if self.stats['processed_files'] % self.intermediate_save_interval == 0:
                self._write_intermediate()

        if self.parallel_processing and len(pdf_files) > 1:
            with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
                fmap = {ex.submit(self._process_one, f): f for f in pdf_files}
                for future in as_completed(fmap):
                    if self.stop_event.is_set():
                        ex.shutdown(wait=False, cancel_futures=True)
                        self.logger.info("Processing stopped by user — saving partial results")
                        break
                    try:
                        record(*future.result())
                    except Exception as e:
                        f = fmap[future]
                        self.logger.error(f"Worker crashed on {f.name}: {e}")
                        with self._lock:
                            self.stats['failed_files'] += 1
        else:
            for pdf_file in pdf_files:
                if self.stop_event.is_set():
                    self.logger.info("Processing stopped by user — saving partial results")
                    break
                try:
                    record(*self._process_one(pdf_file))
                except Exception as e:
                    self.logger.error(f"Error on {pdf_file.name}: {e}")
                    with self._lock:
                        self.stats['failed_files'] += 1

        with self._lock:
            self.stats['total_processing_time'] = time.time() - t0

        self._write_final()
        return self._generate_summary()

    # ── Output writers ───────────────────────────────────────────────────────

    def _screening_headers(self) -> List[str]:
        return ["filename", "decision", "stage", "reasoning", "notes",
                "processing_time", "text_length", "api_tokens_used"]

    def _extraction_headers(self) -> List[str]:
        return ["filename"] + self.extraction_fields + ["processing_time", "api_tokens_used"]

    def _excel_header(self, ws, headers: List[str]):
        hfill  = PatternFill("solid", fgColor="1F3864")
        hfont  = Font(color="FFFFFF", bold=True, size=10)
        halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin   = Side(style='thin', color='C0C0C0')
        bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ci, col in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=col.replace('_', ' ').title())
            cell.fill, cell.font, cell.alignment, cell.border = hfill, hfont, halign, bdr
        ws.row_dimensions[1].height = 28

    def _excel_autofit(self, ws):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width  = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[letter].width = min(width + 3, 65)

    def write_screening_csv(self):
        try:
            with open(self.screening_csv, 'w', newline='', encoding='utf-8') as f:
                w = csv.DictWriter(f, fieldnames=self._screening_headers(), extrasaction='ignore')
                w.writeheader()
                for r in self.screening_results:
                    w.writerow(asdict(r))
        except Exception as e:
            self.logger.error(f"Screening CSV: {e}")

    def write_extraction_csv(self):
        try:
            with open(self.extraction_csv, 'w', newline='', encoding='utf-8') as f:
                w = csv.DictWriter(f, fieldnames=self._extraction_headers(), extrasaction='ignore')
                w.writeheader()
                for r in self.extraction_results:
                    row = {'filename': r.filename,
                           'processing_time': round(r.processing_time, 2),
                           'api_tokens_used': r.api_tokens_used}
                    row.update(r.fields)
                    w.writerow(row)
        except Exception as e:
            self.logger.error(f"Extraction CSV: {e}")

    def write_screening_excel(self):
        try:
            COLOURS = {
                "Likely Include":       "C6EFCE",
                "Likely Exclude":       "FFC7CE",
                "Flag for Review":      "FFEB9C",
                "Flag for Human Review":"DDEBF7",
                "Error":                "EEEEEE",
            }
            alt = PatternFill("solid", fgColor="F5F5F5")
            wb  = Workbook()
            ws  = wb.active
            ws.title = "Screening Results"
            headers  = self._screening_headers()
            self._excel_header(ws, headers)

            for ri, r in enumerate(self.screening_results, 2):
                row = asdict(r)
                for ci, h in enumerate(headers, 1):
                    cell = ws.cell(row=ri, column=ci, value=row.get(h, ""))
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if h == "decision":
                        c = COLOURS.get(str(row.get(h, "")), "FFFFFF")
                        cell.fill = PatternFill("solid", fgColor=c)
                    elif ri % 2 == 0:
                        cell.fill = alt

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            self._excel_autofit(ws)
            wb.save(self.screening_excel)
        except Exception as e:
            self.logger.error(f"Screening Excel: {e}")

    def write_extraction_excel(self):
        try:
            alt     = PatternFill("solid", fgColor="F5F5F5")
            wb      = Workbook()
            ws      = wb.active
            ws.title = "Extracted Data"
            headers  = self._extraction_headers()
            self._excel_header(ws, headers)

            for ri, r in enumerate(self.extraction_results, 2):
                row = {'filename': r.filename,
                       'processing_time': round(r.processing_time, 2),
                       'api_tokens_used': r.api_tokens_used}
                row.update(r.fields)
                for ci, h in enumerate(headers, 1):
                    cell = ws.cell(row=ri, column=ci, value=str(row.get(h, "")))
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if ri % 2 == 0:
                        cell.fill = alt

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            self._excel_autofit(ws)
            wb.save(self.extraction_excel)
        except Exception as e:
            self.logger.error(f"Extraction Excel: {e}")

    def _write_intermediate(self):
        try:
            if self.screening_results:
                self.write_screening_csv()
            if self.extraction_results:
                self.write_extraction_csv()
        except Exception as e:
            self.logger.error(f"Intermediate save: {e}")

    def _write_final(self):
        try:
            if self.screening_results:
                self.write_screening_csv()
                self.write_screening_excel()
            if self.extraction_results:
                self.write_extraction_csv()
                self.write_extraction_excel()
            self.logger.info("Output files written.")
        except Exception as e:
            self.logger.error(f"Final write: {e}")

    # ── Summary ──────────────────────────────────────────────────────────────

    def _generate_summary(self) -> Dict:
        s = self.stats.copy()
        summary = {
            "screening_csv":    str(self.screening_csv),
            "extraction_csv":   str(self.extraction_csv),
            "screening_excel":  str(self.screening_excel),
            "extraction_excel": str(self.extraction_excel),
            "summary_report":   str(self.summary_report),
            "audit_ledger":     str(self.audit_ledger),
            "screened_count":   len(self.screening_results),
            "extracted_count":  len(self.extraction_results),
            "statistics":       s,
            "prisma": {
                "identified": s.get('total_files', 0),
                "screened":   s.get('processed_files', 0),
                "included":   s.get('likely_include', 0),
                "excluded":   s.get('likely_exclude', 0),
                "flagged":    s.get('flag_for_review', 0) + s.get('flag_for_human_review', 0),
                "errors":     s.get('failed_files', 0),
                "extracted":  len(self.extraction_results),
            }
        }

        t = s['total_processing_time']
        n = max(s['processed_files'], 1)
        if t > 0:
            summary['statistics']['files_per_minute']  = round(n / (t / 60), 2)
            summary['statistics']['avg_time_per_file'] = round(t / n, 2)

        try:
            with open(self.summary_report, 'w', encoding='utf-8') as f:
                p = summary['prisma']
                f.write("SYSTEMATIC / SCOPING REVIEW — PROCESSING SUMMARY\n")
                f.write("=" * 55 + "\n\n")
                f.write(f"Completed : {datetime.now():%Y-%m-%d %H:%M:%S}\n")
                f.write(f"Provider  : {self.llm_provider}  Model: {self.llm_model}\n\n")
                f.write("PRISMA FLOW\n")
                f.write(f"  Identified            : {p['identified']}\n")
                f.write(f"  Screened              : {p['screened']}\n")
                f.write(f"  Likely Include        : {p['included']}\n")
                f.write(f"  Likely Exclude        : {p['excluded']}\n")
                f.write(f"  Flagged for Review    : {p['flagged']}\n")
                f.write(f"  Errors                : {p['errors']}\n")
                f.write(f"  Data Extracted        : {p['extracted']}\n\n")
                f.write("PERFORMANCE\n")
                f.write(f"  Total time            : {t:.1f}s\n")
                f.write(f"  Avg per file          : {summary['statistics'].get('avg_time_per_file', 0):.1f}s\n")
                f.write(f"  API tokens used       : {s['total_api_tokens']:,}\n\n")
                f.write("OUTPUT FILES\n")
                f.write(f"  {self.screening_csv}\n")
                f.write(f"  {self.screening_excel}\n")
                if self.extraction_results:
                    f.write(f"  {self.extraction_csv}\n")
                    f.write(f"  {self.extraction_excel}\n")
                f.write(f"  {self.audit_ledger}\n")
        except Exception as e:
            self.logger.error(f"Summary write: {e}")

        return summary


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Universal Systematic / Scoping Review Automation (CLI)",
    )
    parser.add_argument('--api_key',       default="",    help='API key')
    parser.add_argument('--pdf_folder',    required=True, help='Folder with PDFs')
    parser.add_argument('--output_folder', default='output')
    parser.add_argument('--provider',      default='OpenAI')
    parser.add_argument('--model',         default=None)
    parser.add_argument('--base_url',      default=None)
    parser.add_argument('--max_workers',   type=int, default=3)
    parser.add_argument('--rate_limit_delay', type=float, default=1.0)
    parser.add_argument('--disable_cache', action='store_true')
    parser.add_argument('--sequential',    action='store_true')
    parser.add_argument('--two_stage',     action='store_true')
    parser.add_argument('--screening_prompt_file',  default=None)
    parser.add_argument('--extraction_prompt_file', default=None)
    parser.add_argument('--dry_run',       action='store_true')
    args = parser.parse_args()

    if args.dry_run:
        pdfs = list(Path(args.pdf_folder).glob("*.pdf"))
        print(f"Dry-run: {len(pdfs)} PDF(s) in {args.pdf_folder}. No API calls made.")
        return

    kwargs = {}
    if args.base_url:
        kwargs['base_url'] = args.base_url

    sp = Path(args.screening_prompt_file).read_text(encoding='utf-8')  if args.screening_prompt_file  else None
    ep = Path(args.extraction_prompt_file).read_text(encoding='utf-8') if args.extraction_prompt_file else None

    slr = SystematicReviewAutomation(
        api_key=args.api_key,
        pdf_folder=args.pdf_folder,
        output_folder=args.output_folder,
        cache_enabled=not args.disable_cache,
        parallel_processing=not args.sequential,
        max_workers=args.max_workers,
        rate_limit_delay=args.rate_limit_delay,
        screening_prompt=sp,
        extraction_prompt=ep,
        llm_provider=args.provider,
        llm_model=args.model,
        two_stage_screening=args.two_stage,
        **kwargs,
    )

    results = slr.process_pdfs()
    p = results['prisma']
    print(f"\nDone — Screened: {p['screened']}  Included: {p['included']}  "
          f"Excluded: {p['excluded']}  Tokens: {results['statistics']['total_api_tokens']:,}")


if __name__ == "__main__":
    main()
